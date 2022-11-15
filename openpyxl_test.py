#def open():
#     from openpyxl import  load_workbook
#     wb = load_workbook("123.xlsx")
#     #获取工作簿
#     sh1 = wb.active
#     sh2 = wb['Sheet1']
#     sh3 = wb.get_sheet_by_name('Sheet1')
#
#     print(sh1 is sh2 is sh3)
#
#
# if __name__ == '__main__':
#     open()
import openpyxl
from openpyxl import  load_workbook
wb = load_workbook('123.xlsx')
#获取工作簿
sh1 = wb.active
sh2 = wb['Sheet1']
sh3 = wb.get_sheet_by_name('Sheet1')

print(sh1 is sh2 is sh3)





